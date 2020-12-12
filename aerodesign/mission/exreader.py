import openpyxl

def scan_mission_file(path_to_mission_file):
    mission_workbook = openpyxl.load_workbook(path_to_mission_file, read_only = True, data_only = True)
    return mission_workbook.sheetnames


def create_mission_from_sheet(path_to_mission_file,mission_sheet,mission_list):
    mission_workbook = openpyxl.load_workbook(path_to_mission_file, read_only = True, data_only = True)

    id_ = mission_workbook[mission_sheet]['B2']
    name = mission_workbook[mission_sheet]['B3']
    mission_type = mission_workbook[mission_sheet]['B4']
    mission_range = mission_workbook[mission_sheet]['B5']
    payload = mission_workbook[mission_sheet]['B6']