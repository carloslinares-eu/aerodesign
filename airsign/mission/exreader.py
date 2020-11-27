import openpyxl
import airsign.mission.structure


def load_mission_file(path_to_mission_file):
    mission_workbook = openpyxl.load_workbook(path_to_mission_file, read_only=True, data_only=True)
    return mission_workbook


def create_missions(mission_workbook):
    missions = {}
    for mission in mission_workbook.sheetnames:
        mission_sheet = mission_workbook[mission]

        id_ = mission_sheet['B2'].value
        name = mission_sheet['B3'].value
        mission_type = mission_sheet['B4'].value
        mission_range = mission_sheet['B5'].value
        payload = mission_sheet['B6'].value

        missions = {mission: airsign.mission.structure.Mission(id_, name, mission_type, mission_range, payload)}

    return missions


def load_segments(mission_workbook, mission):
    mission_sheet = mission_workbook[mission]
    highest_row = mission_sheet.get_highest_row()

    for row in range(9,  highest_row):
        id_ = mission_sheet.cell(row, 1).value
        name = mission_sheet.cell(row, 2).value
        segment_class = mission_sheet.cell(row, 3).value
        aero_setup = mission_sheet.cell(row, 4).value

        if segment_class == 'Cruise':
            length = mission_sheet.cell(row, 5).value
            speed = mission_sheet.cell(row, 6).value
            segments = {id_:  airsign.mission.structure.Cruise(id_, name, aero_setup, length, speed)}
            break

        elif segment_class == 'Gradient':
            length = mission_sheet.cell(row, 5).value
            speed = mission_sheet.cell(row, 6).value
            climb_gradient = mission_sheet.cell(row, 7).value
            segments = {id_:  airsign.mission.structure.Cruise(id_, name, aero_setup, length, speed, climb_gradient)}
            break

        elif segment_class == 'Loitering':














